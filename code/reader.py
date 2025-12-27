import cv2
import os
import datetime
import yaml
import time
import threading
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Конфигурация путей
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)
DB_PATH = os.path.join(BASE_DIR, "database", "data_user.md")
LOG_DIR = os.path.join(BASE_DIR, "database", "log_kkp")

# Настройки сканирования
SCAN_COOLDOWN = 5  # Задержка между сканированиями в секундах
SCAN_TIMEOUT = 10  # Время ожидания перед следующим сканированием после успешного
DB_RELOAD_INTERVAL = 30  # Интервал перезагрузки базы данных в секундах

class DatabaseManager:
    """Менеджер базы данных с поддержкой авто-обновления"""
    
    def __init__(self, db_path):
        self.db_path = db_path
        self.authorized_users = {}
        self.last_reload_time = 0
        self.db_mtime = 0  # Время последнего изменения файла
        self.lock = threading.Lock()
        self.last_reload_message_time = 0  # Время последнего сообщения о перезагрузке
        self.reload_database()
    
    def reload_database(self, silent=False):
        """Перезагрузка базы данных с проверкой изменений"""
        try:
            # Проверяем время изменения файла
            if not os.path.exists(self.db_path):
                if not silent:
                    print(f"Ошибка: База данных не найдена по пути: {self.db_path}")
                return False
            
            current_mtime = os.path.getmtime(self.db_path)
            if current_mtime <= self.db_mtime:
                return False  # Файл не изменялся
            
            with open(self.db_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Разделяем YAML документы если их несколько
            yaml_documents = content.split('---')
            new_users = {}
            
            for doc in yaml_documents:
                doc = doc.strip()
                if doc:
                    user_data = yaml.safe_load(doc)
                    if user_data and 'ID' in user_data:
                        user_id = user_data['ID']
                        new_users[user_id] = user_data
            
            with self.lock:
                self.authorized_users = new_users
                self.db_mtime = current_mtime
                self.last_reload_time = time.time()
            
            # Выводим сообщение только если прошло достаточно времени с последнего сообщения
            current_time = time.time()
            if current_time - self.last_reload_message_time > 10:  # Не чаще чем раз в 10 секунд
                print(f"База данных перезагружена. Загружено {len(new_users)} пользователей")
                self.last_reload_message_time = current_time
            return True
            
        except Exception as e:
            if not silent:
                print(f"Ошибка перезагрузки базы данных: {e}")
            return False
    
    def get_user(self, user_id):
        """Получение данных пользователя по ID"""
        with self.lock:
            return self.authorized_users.get(user_id)
    
    def get_user_count(self):
        """Получение количества пользователей в базе"""
        with self.lock:
            return len(self.authorized_users)
    
    def should_reload(self):
        """Проверка необходимости перезагрузки базы данных"""
        return time.time() - self.last_reload_time > DB_RELOAD_INTERVAL

def check_access_permission(user_data):
    """Проверка разрешения доступа на основе данных пользователя"""
    if not user_data:
        return False
    
    # Проверяем срок действия
    if 'expiration_date' in user_data:
        expiration = user_data['expiration_date']
        current_date = datetime.datetime.now().date()
        
        try:
            # Если expiration уже является объектом date
            if isinstance(expiration, datetime.date):
                expiration_date = expiration
            else:
                # Если это строка, парсим ее
                expiration_date = datetime.datetime.strptime(str(expiration), '%Y-%m-%d').date()
            
            if current_date > expiration_date:
                print(f"Просроченный пропуск: {expiration_date}")
                return False
                
        except (ValueError, TypeError) as e:
            print(f"Ошибка проверки срока действия: {e}")
            # В случае ошибки парсинга даты, разрешаем доступ для безопасности
            return True
    
    # Проверяем временный пропуск
    if user_data.get('is_temporary', False):
        print("Временный пропуск - требуется дополнительная проверка")
        # Здесь можно добавить дополнительную логику для временных пропусков
    
    return True

def create_excel_log_file(log_file):
    """Создание нового Excel файла с заголовками"""
    try:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Лог доступа"
        
        # Заголовки колонок
        headers = [
            'Время', 
            'ID', 
            'ФИО', 
            'Организация', 
            'Отдел', 
            'Срок действия', 
            'Статус доступа', 
            'Причина'
        ]
        
        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Настраиваем ширину колонок
        column_widths = {
            'A': 20,  # Время
            'B': 15,  # ID
            'C': 30,  # ФИО
            'D': 40,  # Организация
            'E': 30,  # Отдел
            'F': 15,  # Срок действия
            'G': 20,  # Статус доступа
            'H': 30   # Причина
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Сохраняем файл
        workbook.save(log_file)
        print(f"Создан новый лог-файл: {log_file}")
        return True
    except Exception as e:
        print(f"Ошибка создания Excel файла: {e}")
        return False

def log_entry(user_data, access_granted, reason=""):
    """Запись в лог-файл в формате Excel"""
    try:
        today = datetime.datetime.now().strftime("%Y-%m-%d")
        log_file = os.path.join(LOG_DIR, f"log_{today}.xlsx")
        
        # Создаем папку для логов если её нет
        os.makedirs(LOG_DIR, exist_ok=True)
        
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        status = "ДОСТУП РАЗРЕШЕН" if access_granted else "ДОСТУП ЗАПРЕЩЕН"
        
        # Проверяем существование файла
        if not os.path.exists(log_file):
            if not create_excel_log_file(log_file):
                return False
        
        # Загружаем существующий файл
        workbook = load_workbook(log_file)
        worksheet = workbook.active
        
        # Определяем следующую строку для записи
        next_row = worksheet.max_row + 1
        
        # Подготавливаем данные для записи
        if user_data:
            # Форматируем дату для корректного отображения
            expiration = user_data.get('expiration_date', 'Неизвестно')
            if isinstance(expiration, datetime.date):
                expiration = expiration.strftime('%Y-%m-%d')
            
            data = [
                timestamp,
                user_data.get('ID', 'Неизвестно'),
                user_data.get('full_name', 'Неизвестно'),
                user_data.get('organization', 'Неизвестно'),
                user_data.get('department', 'Неизвестно'),
                expiration,
                status,
                reason
            ]
        else:
            data = [
                timestamp,
                'Неизвестно',
                'Неизвестно',
                'Неизвестно',
                'Неизвестно',
                'Неизвестно',
                status,
                reason
            ]
        
        # Записываем данные в строку
        for col, value in enumerate(data, 1):
            worksheet.cell(row=next_row, column=col, value=value)
        
        # Применяем цветовое форматирование
        if access_granted:
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Светло-зеленый
        else:
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Светло-красный
        
        # Применяем цвет ко всей строке
        for col in range(1, 9):  # Колонки от A до H
            worksheet.cell(row=next_row, column=col).fill = fill
        
        # Сохраняем файл
        workbook.save(log_file)
        return True
        
    except Exception as e:
        print(f"Ошибка записи в лог: {e}")
        return False

def process_qr_code(frame):
    """Декодирование QR-кода с использованием OpenCV"""
    try:
        # Инициализируем детектор QR-кодов OpenCV
        qr_detector = cv2.QRCodeDetector()
        
        # Пытаемся обнаружить и декодировать QR-код
        data, bbox, _ = qr_detector.detectAndDecode(frame)
        
        if data and bbox is not None:
            # Визуализируем QR-код на кадре
            if len(bbox) > 0:
                bbox = bbox.astype(int)
                n = len(bbox)
                for i in range(n):
                    cv2.line(frame, tuple(bbox[i][0]), tuple(bbox[(i+1) % n][0]), (0, 255, 0), 3)
            
            return data
        
        return None
    except Exception as e:
        return None

def setup_camera():
    """Настройка и подключение к камере"""
    # Пробуем разные индексы камеры
    for i in range(3):
        cap = cv2.VideoCapture(i)
        
        # Настройки камеры для лучшего распознавания QR
        cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
        cap.set(cv2.CAP_PROP_FPS, 30)
        cap.set(cv2.CAP_PROP_AUTOFOCUS, 1)
        cap.set(cv2.CAP_PROP_FOCUS, 0)
        
        if cap.isOpened():
            # Проверяем, что камера действительно работает
            ret, frame = cap.read()
            if ret and frame is not None:
                print(f"Камера {i} успешно подключена")
                # Даем камере время на инициализацию
                time.sleep(2)
                return cap
            else:
                cap.release()
    
    print("Ошибка: Не удалось подключиться ни к одной камере!")
    return None

def safe_str(value, default="N/A"):
    """Безопасное преобразование значения в строку"""
    if value is None:
        return default
    return str(value)

def display_user_info(frame, user_data, access_granted, countdown=None, db_reload_countdown=None):
    """Отображение информации о пользователе на кадре"""
    color = (0, 255, 0) if access_granted else (0, 0, 255)
    status_text = "ACCESS GRANTED" if access_granted else "ACCESS DENIED"
    
    # Основной статус
    cv2.putText(frame, status_text, (50, 80), 
               cv2.FONT_HERSHEY_SIMPLEX, 1.5, color, 3)
    
    # Отображение обратного отсчета если есть
    if countdown is not None:
        cv2.putText(frame, f"Next scan in: {countdown}s", (50, 250), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
    
    # Отображение времени до обновления БД
    if db_reload_countdown is not None:
        cv2.putText(frame, f"DB reload in: {db_reload_countdown}s", (50, 280), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.5, (200, 200, 0), 1)
    
    if user_data:
        # Форматируем дату для отображения
        expiration = user_data.get('expiration_date', 'N/A')
        if isinstance(expiration, datetime.date):
            expiration = expiration.strftime('%Y-%m-%d')
        
        # Безопасно получаем значения полей
        user_id = safe_str(user_data.get('ID', 'N/A'))
        full_name = safe_str(user_data.get('full_name', 'N/A'))
        org = safe_str(user_data.get('organization', 'N/A'))
        
        # Информация о пользователе
        cv2.putText(frame, f"ID: {user_id}", (50, 120), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2)
        
        # Сокращаем ФИО если слишком длинное
        if len(full_name) > 20:
            full_name = full_name[:20] + "..."
        cv2.putText(frame, f"Name: {full_name}", (50, 150), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 1)
        
        # Сокращаем название организации
        if len(org) > 25:
            org = org[:25] + "..."
        cv2.putText(frame, f"Org: {org}", (50, 180), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 1)
        
        # Срок действия
        exp_color = (0, 255, 0) if access_granted else (0, 165, 255)
        cv2.putText(frame, f"Expires: {expiration}", (50, 210), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.5, exp_color, 1)

def main():
    # Инициализация менеджера базы данных
    db_manager = DatabaseManager(DB_PATH)
    
    # Инициализация камеры
    cap = setup_camera()
    if cap is None:
        return

    print("Система контроля доступа запущена...")
    print("Наведите камеру на QR-код")
    print("Для выхода нажмите 'q'")

    # Состояние системы
    scan_state = "READY"  # READY, PROCESSING, COOLDOWN
    last_scanned_id = None
    scan_start_time = 0
    cooldown_end_time = 0
    current_user_data = None
    current_access_granted = False
    
    # Для отслеживания обновления БД
    last_db_reload_time = time.time()
    db_reload_notified = False  # Флаг для отслеживания уведомления об обновлении

    while True:
        ret, frame = cap.read()
        if not ret:
            print("Ошибка: Не удалось получить кадр с камеры!")
            # Пытаемся переподключиться к камере
            cap.release()
            time.sleep(2)
            cap = setup_camera()
            if cap is None:
                break
            continue

        current_time = time.time()
        
        # Проверка необходимости обновления базы данных
        if db_manager.should_reload():
            if not db_reload_notified:
                print("Обнаружена необходимость обновления базы данных...")
                db_reload_notified = True
            
            db_manager.reload_database(silent=True)  # Тихая перезагрузка без лишних сообщений
            last_db_reload_time = current_time
        else:
            db_reload_notified = False  # Сбрасываем флаг, когда обновление не требуется
        
        # Расчет времени до следующего обновления БД
        db_reload_countdown = max(0, int(DB_RELOAD_INTERVAL - (current_time - db_manager.last_reload_time)))
        
        # Обработка состояний системы
        if scan_state == "READY":
            # Декодирование QR-кода
            qr_data = process_qr_code(frame)
            
            if qr_data:
                scan_state = "PROCESSING"
                scan_start_time = current_time
                last_scanned_id = qr_data
                
                # Поиск пользователя в базе данных (используем актуальные данные)
                current_user_data = db_manager.get_user(qr_data)
                
                if current_user_data:
                    # Проверка дополнительных условий доступа
                    current_access_granted = check_access_permission(current_user_data)
                    
                    if current_access_granted:
                        print(f"✓ ДОСТУП РАЗРЕШЕН - {current_user_data.get('full_name', 'Unknown')}")
                        log_entry(current_user_data, True)
                    else:
                        # Определяем причину отказа
                        expiration = current_user_data.get('expiration_date')
                        if expiration:
                            if isinstance(expiration, datetime.date):
                                current_date = datetime.datetime.now().date()
                                if current_date > expiration:
                                    reason = "Просроченный пропуск"
                                else:
                                    reason = "Доступ ограничен"
                            else:
                                reason = "Проблема с данными пропуска"
                        else:
                            reason = "Доступ ограничен"
                        
                        print(f"✗ ДОСТУП ЗАПРЕЩЕН - {reason}")
                        log_entry(current_user_data, False, reason)
                else:
                    print("✗ ДОСТУП ЗАПРЕЩЕН - Неавторизованный ID")
                    log_entry(None, False, "Неизвестный ID")
                    current_user_data = None
                    current_access_granted = False
        
        elif scan_state == "PROCESSING":
            # Показываем результат сканирования в течение SCAN_COOLDOWN секунд
            if current_time - scan_start_time >= SCAN_COOLDOWN:
                scan_state = "COOLDOWN"
                cooldown_end_time = current_time + SCAN_TIMEOUT
        
        elif scan_state == "COOLDOWN":
            # Ожидание перед следующим сканированием
            if current_time >= cooldown_end_time:
                scan_state = "READY"
                last_scanned_id = None
                current_user_data = None
                print("Готов к сканированию...")

        # Отображение информации в зависимости от состояния
        if scan_state == "READY":
            cv2.putText(frame, "Scan QR Code", (50, 100), 
                       cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 2)
            display_user_info(frame, None, False, None, db_reload_countdown)
        elif scan_state == "PROCESSING":
            remaining = max(0, SCAN_COOLDOWN - (current_time - scan_start_time))
            display_user_info(frame, current_user_data, current_access_granted, int(remaining), db_reload_countdown)
        elif scan_state == "COOLDOWN":
            remaining = max(0, cooldown_end_time - current_time)
            display_user_info(frame, current_user_data, current_access_granted, int(remaining), db_reload_countdown)
            cv2.putText(frame, "Please remove QR code", (50, 310), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)

        # Отображение состояния системы
        state_colors = {"READY": (0, 255, 0), "PROCESSING": (0, 255, 255), "COOLDOWN": (255, 255, 0)}
        cv2.putText(frame, f"State: {scan_state}", (10, 30), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.7, state_colors.get(scan_state, (255, 255, 255)), 2)
        
        # Отображение информации о базе данных
        user_count = db_manager.get_user_count()
        cv2.putText(frame, f"Users in DB: {user_count}", (10, 60), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.5, (200, 200, 200), 1)

        cv2.imshow('u.p.i.c reader', frame)

        # Выход по нажатию 'q'
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()
    print("Система остановлена")

if __name__ == "__main__":
    main()
